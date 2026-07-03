"""伝統的な日本の Excel 設計書フィクスチャの生成器。

実務でよく見る「Excel を方眼紙・帳票レイアウトとして使った設計書」を
openpyxl だけで決定論的に再現する。生成される 5 ファイルはそれぞれ
異なる構造化の難所を代表する:

  screen_item_def.xlsx  画面項目定義書  ヘッダブロック(ラベル:値) + 2段結合ヘッダ表
  table_def.xlsx        テーブル定義書  複数シート + ○×印 + 空欄混じりの定義表
  houganshi_spec.xlsx   機能仕様書      Excel 方眼紙 (1文字幅列 + 横結合セルの文章)
  test_spec.xlsx        試験項目書      縦結合の分類列 + セル内改行の手順
  basic_design.xlsx     基本設計書      表紙/改訂履歴/本体の複数シート + 1シート複数表
  network_diagram.xlsx  ネットワーク構成図  オートシェイプ(ノード)+コネクタ(接続線) の図形

正解データは truth/<名前>.json に手書きで宣言する (このスクリプトからは
生成しない)。生成データを変えたら truth も必ず追随させること。

使い方::

    python make_fixtures.py [出力ディレクトリ]   # 既定は ./fixtures
"""

from __future__ import annotations

import os
import shutil
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 見た目 (抽出には影響しないが、人が Excel で開いて確認できるようにする) ──
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
_TITLE_FONT = Font(size=14, bold=True)
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")


def _put(ws, ref: str, value, *, fill=None, font=None, align=None, border=True):
    """ref ("B4" または "B4:D4") に値を書き、結合と体裁を適用する。"""
    top_left = ref.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    if ":" in ref:
        ws.merge_cells(ref)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = _BORDER
    return cell


def _table(ws, first_row: int, header: list[str], rows: list[list], *, first_col=1):
    """1段ヘッダの素朴な表を書く。"""
    for j, h in enumerate(header):
        c = ws.cell(row=first_row, column=first_col + j, value=h)
        c.fill = _HEADER_FILL
        c.font = _BOLD
        c.border = _BORDER
    for i, row in enumerate(rows, start=first_row + 1):
        for j, v in enumerate(row):
            c = ws.cell(row=i, column=first_col + j, value=v)
            c.border = _BORDER
            c.alignment = _WRAP


# ── 1. 画面項目定義書: ラベル:値ヘッダブロック + 2段結合ヘッダの項目表 ──
def build_screen_item_def(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "画面項目定義"
    for col, w in zip("ABCDEFGH", (6, 16, 16, 10, 6, 6, 14, 28)):
        ws.column_dimensions[col].width = w

    _put(ws, "A1:H1", "画面項目定義書", font=_TITLE_FONT, align=_CENTER)

    _put(ws, "A3", "システム名", fill=_LABEL_FILL)
    _put(ws, "B3:D3", "受注管理システム")
    _put(ws, "A4", "画面ID", fill=_LABEL_FILL)
    _put(ws, "B4", "SCR-001")
    _put(ws, "C4", "画面名", fill=_LABEL_FILL)
    _put(ws, "D4:E4", "受注登録画面")
    _put(ws, "A5", "作成者", fill=_LABEL_FILL)
    _put(ws, "B5", "山田太郎")
    _put(ws, "C5", "作成日", fill=_LABEL_FILL)
    _put(ws, "D5", "2026/04/01")
    _put(ws, "A6", "版数", fill=_LABEL_FILL)
    _put(ws, "B6", "1.2")

    # 2段ヘッダ: 「属性」の下に 型/桁/必須 をぶら下げる伝統形式
    for ref, text in [
        ("A8:A9", "No"),
        ("B8:B9", "項目名"),
        ("C8:C9", "物理名"),
        ("D8:F8", "属性"),
        ("G8:G9", "初期値"),
        ("H8:H9", "備考"),
    ]:
        _put(ws, ref, text, fill=_HEADER_FILL, font=_BOLD, align=_CENTER)
    for ref, text in [("D9", "型"), ("E9", "桁"), ("F9", "必須")]:
        _put(ws, ref, text, fill=_HEADER_FILL, font=_BOLD, align=_CENTER)

    items = [
        [1, "受注番号", "ORDER_NO", "文字列", 10, "○", "自動採番", "主キー"],
        [2, "受注日", "ORDER_DATE", "日付", None, "○", "当日日付", None],
        [3, "顧客コード", "CUST_CODE", "文字列", 8, "○", None, "顧客マスタ参照"],
        [4, "顧客名", "CUST_NAME", "文字列", 40, None, None, "顧客コードから自動表示"],
        [5, "受注金額", "ORDER_AMOUNT", "数値", 12, "○", 0, "税込金額"],
        [6, "備考", "REMARKS", "文字列", 200, None, None, None],
    ]
    for i, row in enumerate(items, start=10):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = _BORDER
    wb.save(str(path))


# ── 2. テーブル定義書: 一覧シート + 定義シート (○×印・空欄混じり) ──
def build_table_def(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "テーブル一覧"
    _put(ws, "A1:E1", "テーブル定義書", font=_TITLE_FONT, align=_CENTER)
    _table(
        ws,
        3,
        ["No", "テーブルID", "論理名", "物理名", "備考"],
        [
            [1, "TBL-001", "受注", "T_ORDER", "受注ヘッダ情報"],
            [2, "TBL-002", "受注明細", "T_ORDER_DETAIL", "受注の明細行"],
        ],
    )

    ws2 = wb.create_sheet("TBL-001_受注")
    _put(ws2, "A1", "テーブルID", fill=_LABEL_FILL)
    _put(ws2, "B1", "TBL-001")
    _put(ws2, "C1", "論理名", fill=_LABEL_FILL)
    _put(ws2, "D1", "受注")
    _put(ws2, "E1", "物理名", fill=_LABEL_FILL)
    _put(ws2, "F1", "T_ORDER")
    _table(
        ws2,
        3,
        ["No", "論理名", "物理名", "型", "桁", "NULL", "PK", "備考"],
        [
            [1, "受注番号", "ORDER_NO", "CHAR", 10, "×", "○", "主キー"],
            [2, "受注日", "ORDER_DATE", "DATE", None, "×", None, None],
            [3, "顧客コード", "CUST_CODE", "CHAR", 8, "×", None, "顧客マスタFK"],
            [4, "受注金額", "ORDER_AMOUNT", "NUMBER", 12, "○", None, "税込"],
            [5, "更新日時", "UPDATED_AT", "TIMESTAMP", None, "×", None, None],
        ],
    )
    wb.save(str(path))


# ── 3. 機能仕様書: Excel 方眼紙 (1文字幅の列 + 横結合セルに文章を流す) ──
def build_houganshi_spec(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "機能仕様"
    # A..AN の 40 列を 1 文字幅にして「方眼紙」を作る
    for idx in range(1, 41):
        col = ""
        n = idx
        while n:
            n, rem = divmod(n - 1, 26)
            col = chr(65 + rem) + col
        ws.column_dimensions[col].width = 2.5

    _put(ws, "B2:AM2", "機能仕様書　受注登録機能", font=_TITLE_FONT, align=_CENTER)

    _put(ws, "B4:H4", "1. 概要", font=_BOLD, border=False)
    _put(
        ws,
        "C6:AL8",
        "本機能は、営業担当者が顧客からの受注情報を登録するための機能である。\n"
        "登録された受注情報は受注テーブルに保存され、出荷管理機能から参照される。",
        align=_WRAP,
    )

    _put(ws, "B10:H10", "2. 処理概要", font=_BOLD, border=False)
    steps = [
        ("C12:AL12", "(1) 画面から入力された受注情報の妥当性チェックを行う。"),
        ("C14:AL14", "(2) チェックエラーがある場合はエラーメッセージを表示し、処理を中断する。"),
        ("C16:AL16", "(3) チェックOKの場合、受注テーブルに受注情報を登録する。"),
        ("C18:AL18", "(4) 登録完了後、受注番号を画面に表示する。"),
    ]
    for ref, text in steps:
        _put(ws, ref, text, border=False)

    _put(ws, "B20:H20", "3. 制約事項", font=_BOLD, border=False)
    _put(ws, "C22:AL23", "受注金額が100万円を超える場合は、上長承認が必要となる。", align=_WRAP)
    wb.save(str(path))


# ── 4. 試験項目書: 大分類/中分類の縦結合 + セル内改行の手順 ──
def build_test_spec(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "試験項目"
    for col, w in zip("ABCDEFGH", (6, 12, 12, 30, 34, 34, 8, 10)):
        ws.column_dimensions[col].width = w

    _put(ws, "A1:H1", "単体試験項目書　受注登録画面", font=_TITLE_FONT, align=_CENTER)

    header = ["No", "大分類", "中分類", "試験内容", "試験手順", "期待結果", "結果", "確認者"]
    for j, h in enumerate(header, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = _HEADER_FILL
        c.font = _BOLD
        c.border = _BORDER

    rows = [
        [1, "画面表示", "初期表示", "画面初期表示時の項目状態を確認する",
         "1. メニューから受注登録を選択する\n2. 画面表示を確認する",
         "全項目が初期値で表示されること", "OK", "佐藤"],
        [2, None, None, "受注番号が自動採番されること",
         "1. 画面を表示する\n2. 受注番号欄を確認する",
         "受注番号欄に新規番号が表示されること", "OK", "佐藤"],
        [3, "入力チェック", "必須チェック", "顧客コード未入力でエラーとなること",
         "1. 顧客コードを空にする\n2. 登録ボタンを押下する",
         "「顧客コードを入力してください」と表示されること", "NG", "鈴木"],
        [4, None, None, "受注日未入力でエラーとなること",
         "1. 受注日を空にする\n2. 登録ボタンを押下する",
         "「受注日を入力してください」と表示されること", "OK", "鈴木"],
        [5, None, "形式チェック", "受注金額に文字列を入力するとエラーとなること",
         "1. 受注金額に「あああ」を入力する\n2. 登録ボタンを押下する",
         "「受注金額は数値で入力してください」と表示されること", "OK", "鈴木"],
        [6, "登録処理", "正常系", "正常値の入力で受注が登録されること",
         "1. 全項目に正常値を入力する\n2. 登録ボタンを押下する",
         "受注テーブルに1件登録され、完了メッセージが表示されること", "未実施", None],
    ]
    for i, row in enumerate(rows, start=4):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = _BORDER
            c.alignment = _WRAP
    # 分類列の縦結合 (伝統の「同上はセル結合」)
    ws.merge_cells("B4:B5")   # 画面表示
    ws.merge_cells("C4:C5")   # 初期表示
    ws.merge_cells("B6:B8")   # 入力チェック
    ws.merge_cells("C6:C7")   # 必須チェック
    wb.save(str(path))


# ── 5. 基本設計書: 表紙 + 改訂履歴 + 本体 (1 シートに複数の表) ──
def build_basic_design(path: Path) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "表紙"
    _put(cover, "C8:J10", "受注管理システム", font=Font(size=20, bold=True),
         align=_CENTER, border=False)
    _put(cover, "C12:J13", "基本設計書", font=Font(size=24, bold=True),
         align=_CENTER, border=False)
    _put(cover, "C16:J16", "第1.2版", align=_CENTER, border=False)
    _put(cover, "C18:J18", "2026年4月1日", align=_CENTER, border=False)
    _put(cover, "C20:J20", "株式会社サンプルシステムズ", align=_CENTER, border=False)

    hist = wb.create_sheet("改訂履歴")
    _put(hist, "A1", "改訂履歴", font=_TITLE_FONT, border=False)
    _table(
        hist,
        3,
        ["版数", "改訂日", "改訂内容", "担当"],
        [
            ["1.0", "2026/01/15", "初版作成", "山田"],
            ["1.1", "2026/02/20", "受注金額の桁数を12桁に変更", "山田"],
            ["1.2", "2026/04/01", "上長承認フローの追記", "佐藤"],
        ],
    )

    body = wb.create_sheet("処理設計")
    _put(body, "A1", "5. 処理設計", font=_TITLE_FONT, border=False)
    _put(body, "A3:J4", "受注登録機能の処理一覧と外部インターフェースを以下に示す。",
         align=_WRAP, border=False)
    _put(body, "A6", "5.1 処理一覧", font=_BOLD, border=False)
    _table(
        body,
        7,
        ["No", "処理ID", "処理名", "処理概要"],
        [
            [1, "P-001", "受注登録", "受注情報を受注テーブルに登録する"],
            [2, "P-002", "受注検索", "条件に一致する受注情報を検索する"],
            [3, "P-003", "受注取消", "受注情報を論理削除する"],
        ],
    )
    # 2 行以上の空白ギャップ -> 抽出器は別の表として分割するはず
    _put(body, "A13", "5.2 外部インターフェース一覧", font=_BOLD, border=False)
    _table(
        body,
        14,
        ["No", "IF-ID", "接続先", "方式", "概要"],
        [
            [1, "IF-001", "出荷管理システム", "ファイル連携", "受注確定データを日次で連携する"],
            [2, "IF-002", "会計システム", "API連携", "売上計上データをリアルタイム連携する"],
        ],
    )
    wb.save(str(path))


# ── 6. ネットワーク構成図: 図形 (ノード) + コネクタ (接続線) で描く構成図 ──
#    openpyxl は autoshape を書けないので、保存後に drawing XML を直接注入する。
_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_AMAIN = "http://schemas.openxmlformats.org/drawingml/2006/main"
_RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _sp_xml(idx: int, node: dict) -> str:
    """1 個のノード (テキスト入り矩形) の twoCellAnchor XML を組み立てる。"""
    col, row = node["col"], node["row"]
    paras = "".join(
        f"<a:p><a:r><a:t>{line}</a:t></a:r></a:p>" for line in node["text"].split("\n")
    )
    return (
        "<xdr:twoCellAnchor>"
        f"<xdr:from><xdr:col>{col}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>{col + 2}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row + 3}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f'<xdr:sp macro="" textlink=""><xdr:nvSpPr>'
        f'<xdr:cNvPr id="{idx}" name="{node["name"]}"/><xdr:cNvSpPr/></xdr:nvSpPr>'
        f'<xdr:spPr><a:prstGeom prst="{node.get("geom", "rect")}"><a:avLst/>'
        "</a:prstGeom></xdr:spPr>"
        f"<xdr:txBody><a:bodyPr/>{paras}</xdr:txBody></xdr:sp>"
        "<xdr:clientData/></xdr:twoCellAnchor>"
    )


def _cxn_xml(idx: int, frm: tuple[int, int], to: tuple[int, int]) -> str:
    """1 本のコネクタ (テキスト無し接続線) の twoCellAnchor XML。抽出対象外の確認用。"""
    return (
        "<xdr:twoCellAnchor>"
        f"<xdr:from><xdr:col>{frm[0]}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{frm[1]}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>{to[0]}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{to[1]}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f'<xdr:cxnSp macro=""><xdr:nvCxnSpPr>'
        f'<xdr:cNvPr id="{idx}" name="Conn-{idx}"/><xdr:cNvCxnSpPr/></xdr:nvCxnSpPr>'
        '<xdr:spPr><a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>'
        "</xdr:spPr></xdr:cxnSp>"
        "<xdr:clientData/></xdr:twoCellAnchor>"
    )


def _inject_drawing(path: Path, sheet_no: int, anchors_xml: str) -> None:
    """保存済み xlsx の sheetN に drawing (figures) を注入して再パッケージする。"""
    tmpd = path.parent / (path.stem + "_unzip")
    if tmpd.exists():
        shutil.rmtree(tmpd)
    with zipfile.ZipFile(path) as z:
        z.extractall(tmpd)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{_XDR}" xmlns:a="{_AMAIN}">{anchors_xml}</xdr:wsDr>'
    )
    (tmpd / "xl" / "drawings").mkdir(parents=True, exist_ok=True)
    (tmpd / "xl" / "drawings" / "drawing1.xml").write_text(drawing, encoding="utf-8")

    sheet_path = tmpd / "xl" / "worksheets" / f"sheet{sheet_no}.xml"
    s = sheet_path.read_text(encoding="utf-8")
    if "xmlns:r=" not in s:
        s = s.replace("<worksheet ", f'<worksheet xmlns:r="{_RNS}" ', 1)
    s = s.replace("</worksheet>", '<drawing r:id="rIdDraw"/></worksheet>')
    sheet_path.write_text(s, encoding="utf-8")

    rels_dir = tmpd / "xl" / "worksheets" / "_rels"
    rels_dir.mkdir(parents=True, exist_ok=True)
    rel = (
        f'<Relationship Id="rIdDraw" Type="{_RNS}/drawing" '
        'Target="../drawings/drawing1.xml"/>'
    )
    rels_path = rels_dir / f"sheet{sheet_no}.xml.rels"
    if rels_path.exists():
        r = rels_path.read_text(encoding="utf-8").replace(
            "</Relationships>", rel + "</Relationships>"
        )
    else:
        r = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/'
            f'package/2006/relationships">{rel}</Relationships>'
        )
    rels_path.write_text(r, encoding="utf-8")

    ct_path = tmpd / "[Content_Types].xml"
    ct = ct_path.read_text(encoding="utf-8")
    override = (
        '<Override PartName="/xl/drawings/drawing1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
    )
    if "/xl/drawings/drawing1.xml" not in ct:
        ct = ct.replace("</Types>", override + "</Types>")
    ct_path.write_text(ct, encoding="utf-8")

    path.unlink()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(tmpd):
            for f in files:
                fp = Path(root) / f
                z.write(fp, fp.relative_to(tmpd).as_posix())
    shutil.rmtree(tmpd)


def build_network_diagram(path: Path) -> None:
    """図形で描いたネットワーク構成図。ノード名・IP は図形テキストとして救出される。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ネットワーク構成図"
    _put(ws, "A1", "システムネットワーク構成図", font=_TITLE_FONT, border=False)
    _put(ws, "A2", "作成日: 2026/04/01", border=False)
    # 図の下に機器一覧表を置き、表と図形が同居しても両方取れることを見る
    _table(
        ws,
        20,
        ["機器名", "IPアドレス", "役割"],
        [
            ["FWサーバ", "192.168.0.1", "ファイアウォール"],
            ["Webサーバ", "192.168.1.10", "リバースプロキシ"],
            ["APサーバ", "192.168.2.20", "業務ロジック"],
            ["DBサーバ", "192.168.3.30", "データ永続化"],
        ],
    )
    wb.save(str(path))

    # ノード (col,row は 0 始まり)。横一列に FW→Web→AP→DB を並べる。
    nodes = [
        {"name": "Node-FW", "text": "FWサーバ\n192.168.0.1", "col": 1, "row": 4, "geom": "rect"},
        {"name": "Node-Web", "text": "Webサーバ (DMZ)\n192.168.1.10", "col": 5, "row": 4, "geom": "roundRect"},
        {"name": "Node-AP", "text": "APサーバ\n192.168.2.20", "col": 9, "row": 4, "geom": "rect"},
        {"name": "Node-DB", "text": "DBサーバ\n192.168.3.30", "col": 13, "row": 4, "geom": "rect"},
    ]
    anchors = [_sp_xml(i + 2, n) for i, n in enumerate(nodes)]
    # コネクタ (隣接ノード間)。テキストが無いので抽出には現れないはず。
    conns = [((3, 5), (5, 5)), ((7, 5), (9, 5)), ((11, 5), (13, 5))]
    anchors += [_cxn_xml(100 + i, frm, to) for i, (frm, to) in enumerate(conns)]
    _inject_drawing(path, sheet_no=1, anchors_xml="".join(anchors))


BUILDERS = {
    "screen_item_def.xlsx": build_screen_item_def,
    "table_def.xlsx": build_table_def,
    "houganshi_spec.xlsx": build_houganshi_spec,
    "test_spec.xlsx": build_test_spec,
    "basic_design.xlsx": build_basic_design,
    "network_diagram.xlsx": build_network_diagram,
}


def build_all(out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for name, builder in BUILDERS.items():
        p = out_dir / name
        builder(p)
        paths.append(p)
    return paths


def main(argv: list[str]) -> int:
    out_dir = Path(argv[1]) if len(argv) > 1 else Path(__file__).resolve().parent / "fixtures"
    for p in build_all(out_dir):
        print(p)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
