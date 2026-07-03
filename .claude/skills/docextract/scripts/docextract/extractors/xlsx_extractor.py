"""Excel (.xlsx) の抽出器。

各シートから「表領域」を検出して抽出する。

Excel の使用範囲 (used range) は書式だけのセルや削除済みデータの残骸で
実データより大きく膨らむことが多く、そのまま 2 次元配列にすると大量の
空文字が混じる。そこで非空セルの連結成分 (connected components) から
表領域を検出し、領域ごとに 1 つの表として抽出する:

- 空白ギャップ 1 行/列以内で隣接するセルは同じ表とみなす
  (表中のスペーサー行で分断されないため)
- 2 行/列以上の空白で離れたセル群は別の表として分割する
- 各表の内部でも完全な空行・空列は除去する
- 孤立した単一セル (1x1 の領域) は表ではなくテキスト要素として出す
  (表紙のタイトルや注記が「1x1 の表」で大量に出るのを防ぐ)

結合セルは openpyxl では左上セルにのみ値が入るため、幅 1 の縦結合
(表の分類列で「同上」を表す慣習) に限り値を結合範囲の全行へ展開する。
横結合・面結合はタイトルや文章のレイアウト用途が大半で、展開すると
同じ値が重複するだけなので左上セルのまま維持する。

シートに埋め込まれた画像も取り出す。

さらに、オートシェイプ・テキストボックス (ネットワーク構成図やフロー図の
「ノード」) の中のテキストも救出する。日本の設計書では構成図・フロー図を
Excel 図形で描く慣習が根強いが、openpyxl は図形 (drawing の <xdr:sp>) を
読まないためセル外のノード名・IP 等が丸ごと落ちる。ここでは xlsx (zip) 内の
xl/drawings/drawingN.xml を直接パースし、図形テキストを TextElement として
出力する。

コネクタ (<xdr:cxnSp>) が表す接続関係 (トポロジ) も復元する。コネクタが
図形に「接着」されていれば端点は <a:stCxn>/<a:endCxn> に接続先シェイプの
id で記録されるのでそれを使い、無ければコネクタ端点セルに最も近いノードを
幾何的に対応付ける。復元したエッジは「接続元/接続先」の 2 列テーブル
(location に kind="diagram_topology") として 1 シートにつき 1 つ出力する。
"""

from __future__ import annotations

import posixpath
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from ..models import ExtractionResult, ImageElement, TableElement, TextElement
from .base import ImageSaver

# 同じ表とみなす空白ギャップの許容幅 (行/列数)
_GAP = 1

# コネクタ端点をノードに幾何的にスナップする最大距離 (セル数、マンハッタン)。
# これを超えて離れた端点はどのノードにも接続していないとみなし棄却する。
_CXN_SNAP = 2

# OOXML 名前空間
_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}
_REL_DRAWING = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"


def extract_xlsx(path: Path, saver: ImageSaver) -> ExtractionResult:
    # data_only=True で数式ではなくキャッシュされた計算結果を読む
    wb = load_workbook(str(path), data_only=True)
    result = ExtractionResult(source=path.name, file_type="xlsx")

    props = wb.properties
    result.metadata = {
        "title": props.title or None,
        "author": props.creator or None,
        "created": props.created.isoformat() if props.created else None,
        "modified": props.modified.isoformat() if props.modified else None,
        "sheets": wb.sheetnames,
    }

    drawings_by_sheet = _extract_drawing_shapes(path, result)

    for ws in wb.worksheets:
        grid = _sheet_to_grid(ws)
        for top, left, bottom, right in _find_table_regions(grid):
            rows = _region_to_rows(grid, top, left, bottom, right)
            if len(rows) == 1 and len(rows[0]) == 1:
                # 孤立セルはタイトル・ラベル・注記であって表ではない
                result.elements.append(
                    TextElement(
                        content=rows[0][0],
                        location={
                            "sheet": ws.title,
                            "cell": f"{_col_letter(left + 1)}{top + 1}",
                        },
                    )
                )
                continue
            cell_range = (
                f"{_col_letter(left + 1)}{top + 1}:"
                f"{_col_letter(right + 1)}{bottom + 1}"
            )
            result.elements.append(
                TableElement(
                    rows=rows,
                    location={"sheet": ws.title, "range": cell_range},
                )
            )
        for img in getattr(ws, "_images", []):
            try:
                data = img._data()
            except Exception:
                continue
            ext = getattr(img, "format", None) or "png"
            rel_path = saver.save(data, str(ext))
            result.elements.append(
                ImageElement(
                    file=rel_path,
                    format=str(ext).lower(),
                    location={"sheet": ws.title, "anchor": _anchor_cell(img)},
                )
            )
        drawing = drawings_by_sheet.get(ws.title)
        if not drawing:
            continue
        for node in drawing["nodes"]:
            loc: dict[str, str] = {"sheet": ws.title}
            if node["cell"]:
                loc["cell"] = node["cell"]
            if node["name"]:
                loc["shape_name"] = node["name"]
            if node["id"]:
                loc["shape_id"] = node["id"]
            result.elements.append(
                TextElement(content=node["text"], style="shape", location=loc)
            )
        if drawing["edges"]:
            result.elements.append(
                TableElement(
                    rows=[["接続元", "接続先"]]
                    + [[e["src"], e["dst"]] for e in drawing["edges"]],
                    location={"sheet": ws.title, "kind": "diagram_topology"},
                )
            )
    return result


def _cell_to_str(v) -> str:
    """セル値を文字列化する。None と空白のみの値は空とみなす。"""
    if v is None:
        return ""
    s = str(v)
    return s if s.strip() else ""


def _sheet_to_grid(ws) -> list[list[str]]:
    grid = [[_cell_to_str(v) for v in row] for row in ws.iter_rows(values_only=True)]
    _fill_vertical_merges(ws, grid)
    return grid


def _fill_vertical_merges(ws, grid: list[list[str]]) -> None:
    """幅 1 の縦結合セルの値を結合範囲の全行へ展開する (「同上」の復元)。

    横結合・面結合 (幅 2 以上) はタイトル・文章のレイアウト用途が大半で、
    展開しても同じ値が重複するだけなので対象外 (左上セルのまま)。
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_col != rng.max_col or rng.min_row == rng.max_row:
            continue
        r0, c = rng.min_row - 1, rng.min_col - 1
        if r0 >= len(grid) or c >= len(grid[r0]):
            continue
        value = grid[r0][c]
        if not value:
            continue
        for r in range(r0 + 1, min(rng.max_row, len(grid))):
            grid[r][c] = value


def _find_table_regions(
    grid: list[list[str]],
) -> list[tuple[int, int, int, int]]:
    """非空セルの連結成分から表領域の外接矩形を求める。

    空白ギャップ _GAP 以内 (チェビシェフ距離 _GAP+1 以内) のセル同士を
    連結とみなして flood fill し、成分ごとの外接矩形を返す。
    戻り値は 0 始まりの (top, left, bottom, right) を位置順に並べたもの。
    """
    filled = {
        (r, c)
        for r, row in enumerate(grid)
        for c, v in enumerate(row)
        if v
    }
    reach = _GAP + 1
    offsets = [
        (dr, dc)
        for dr in range(-reach, reach + 1)
        for dc in range(-reach, reach + 1)
        if dr or dc
    ]
    regions: list[tuple[int, int, int, int]] = []
    seen: set[tuple[int, int]] = set()
    for start in filled:
        if start in seen:
            continue
        seen.add(start)
        stack = [start]
        top, left = start
        bottom, right = start
        while stack:
            r, c = stack.pop()
            top = min(top, r)
            bottom = max(bottom, r)
            left = min(left, c)
            right = max(right, c)
            for dr, dc in offsets:
                nb = (r + dr, c + dc)
                if nb in filled and nb not in seen:
                    seen.add(nb)
                    stack.append(nb)
        regions.append((top, left, bottom, right))
    return sorted(regions)


def _region_to_rows(
    grid: list[list[str]], top: int, left: int, bottom: int, right: int
) -> list[list[str]]:
    """外接矩形から行列を切り出し、内部の完全な空行・空列を除去する。"""
    rows = [row[left : right + 1] for row in grid[top : bottom + 1]]
    rows = [r for r in rows if any(r)]
    keep = [j for j in range(right - left + 1) if any(r[j] for r in rows)]
    return [[r[j] for j in keep] for r in rows]


def _col_letter(n: int) -> str:
    """1 始まりの列番号を A1 形式の列文字に変換する。"""
    col = ""
    while n:
        n, rem = divmod(n - 1, 26)
        col = chr(65 + rem) + col
    return col


def _anchor_cell(img) -> str | None:
    try:
        anc = img.anchor._from
        return f"{_col_letter(anc.col + 1)}{anc.row + 1}"
    except Exception:
        return None


def _extract_drawing_shapes(
    path: Path, result: ExtractionResult
) -> dict[str, dict]:
    """図形のノード (テキスト) と接続 (トポロジ) をシートごとに取り出す。

    openpyxl は drawing の <xdr:sp>/<xdr:cxnSp> を読まないため、xlsx (zip) 内の
    xl/drawings/drawingN.xml を直接パースする。戻り値は
    {シート名: {"nodes": [dict, ...], "edges": [dict, ...]}}。
    nodes の各要素は {"id","name","text","cell"}、edges は {"src","dst","name"}。
    """
    drawings: dict[str, dict] = {}
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            for sheet_title, sheet_part in _worksheet_parts(zf, names):
                drawing_part = _drawing_part_for_sheet(zf, sheet_part, names)
                if not drawing_part or drawing_part not in names:
                    continue
                try:
                    root = ET.fromstring(zf.read(drawing_part))
                except ET.ParseError as e:
                    result.note_degraded(
                        "xlsx-drawing",
                        f"drawing のパースに失敗: {e}",
                        sheet=sheet_title,
                        part=drawing_part,
                    )
                    continue
                nodes, edges = _parse_drawing(root)
                if nodes or edges:
                    bucket = drawings.setdefault(
                        sheet_title, {"nodes": [], "edges": []}
                    )
                    bucket["nodes"].extend(nodes)
                    bucket["edges"].extend(edges)
    except (zipfile.BadZipFile, OSError):
        # zip として開けない (壊れている等) 場合は図形抽出のみ諦める。
        # セル・画像側の抽出は既に完了しているので握り潰さず記録する。
        result.note_degraded("xlsx-drawing", "drawing 部の読み取りに失敗")
    return drawings


def _worksheet_parts(zf, names: set[str]):
    """(シート名, ワークシート部のパス) を workbook.xml のシート順で返す。"""
    if "xl/workbook.xml" not in names or "xl/_rels/workbook.xml.rels" not in names:
        return
    rels = _rels_map(zf.read("xl/_rels/workbook.xml.rels"))
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    sheets = wb.find("{%s}sheets" % _NS_MAIN)
    if sheets is None:
        return
    for sheet in sheets.findall("{%s}sheet" % _NS_MAIN):
        name = sheet.get("name")
        rid = sheet.get("{%s}id" % _NS["r"])
        target = rels.get(rid)
        if not name or not target:
            continue
        # workbook.xml (xl/ 配下) 相対の Target を zip 内パスへ正規化。
        yield name, _resolve_part("xl", target)


def _drawing_part_for_sheet(zf, sheet_part: str, names: set[str]) -> str | None:
    """ワークシート部に紐づく drawing 部のパスを rels から解決する。"""
    rels_part = posixpath.join(
        posixpath.dirname(sheet_part), "_rels", posixpath.basename(sheet_part) + ".rels"
    )
    if rels_part not in names:
        return None
    for rid, target in _rels_map(zf.read(rels_part), by_type=_REL_DRAWING).items():
        # sheet 部のディレクトリ基準で Target を正規化する (../drawings/..)。
        return _resolve_part(posixpath.dirname(sheet_part), target)
    return None


def _resolve_part(base_dir: str, target: str) -> str:
    """.rels の Target を zip 内パス (先頭スラッシュ無し) へ正規化する。

    先頭が "/" の Target はパッケージ (zip) ルート基準の絶対参照なので
    先頭スラッシュを外す。それ以外は base_dir 相対として解決する。
    """
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target)).lstrip("/")


def _rels_map(data: bytes, by_type: str | None = None) -> dict[str, str]:
    """.rels を {Id: Target} に。by_type 指定時は Type 一致の関係のみ。"""
    root = ET.fromstring(data)
    out: dict[str, str] = {}
    for rel in root.findall("{%s}Relationship" % _NS["pr"]):
        if by_type is not None and rel.get("Type") != by_type:
            continue
        rid, target = rel.get("Id"), rel.get("Target")
        if rid and target:
            out[rid] = target
    return out


def _parse_drawing(root) -> tuple[list[dict], list[dict]]:
    """drawing XML からノード (テキスト図形) と接続 (コネクタ由来) を組み立てる。

    テキストのある <xdr:sp> をノードとし、<xdr:cxnSp> の各端点を接続先ノードに
    解決してエッジにする。端点は明示接続 (<a:stCxn>/<a:endCxn> の id) を最優先し、
    無ければ端点セルに最も近いノードへ幾何的に対応付ける。グループ (<xdr:grpSp>)
    内の図形・コネクタも子孫探索で拾う。
    """
    sp_tag = "{%s}sp" % _NS["xdr"]
    cxn_tag = "{%s}cxnSp" % _NS["xdr"]

    nodes: list[dict] = []
    node_by_id: dict[str, dict] = {}
    connectors: list[dict] = []

    for anchor in root:
        rect = _anchor_rect(anchor)  # (col0,row0,col1,row1) セル座標
        cell = _rect_cell(rect)
        for sp in anchor.iter(sp_tag):
            text = _shape_text(sp)
            if not text:
                continue
            node = {
                "id": _shape_id(sp),
                "name": _shape_name(sp),
                "text": text,
                "cell": cell,
                "rect": rect,
            }
            nodes.append(node)
            if node["id"]:
                node_by_id[node["id"]] = node
        for cxn in anchor.iter(cxn_tag):
            connectors.append(
                {
                    "name": _shape_name(cxn),
                    "st": _cxn_end_id(cxn, "stCxn"),
                    "end": _cxn_end_id(cxn, "endCxn"),
                    "rect": rect,
                }
            )

    edges: list[dict] = []
    for c in connectors:
        src = _resolve_endpoint(c["st"], (c["rect"][0], c["rect"][1]), node_by_id, nodes)
        dst = _resolve_endpoint(c["end"], (c["rect"][2], c["rect"][3]), node_by_id, nodes)
        if src is None or dst is None or src is dst:
            continue
        edges.append(
            {"src": _node_label(src), "dst": _node_label(dst), "name": c["name"]}
        )
    # 公開しない内部キー (rect) を落とす。
    for n in nodes:
        n.pop("rect", None)
    return nodes, edges


def _node_label(node: dict) -> str:
    """エッジ表示用のノード名。図形テキスト先頭行、無ければ図形名。"""
    first = node["text"].splitlines()[0].strip() if node["text"] else ""
    return first or node["name"]


def _resolve_endpoint(
    cxn_id: str | None,
    point: tuple[int, int],
    node_by_id: dict[str, dict],
    nodes: list[dict],
) -> dict | None:
    """コネクタ端点を接続先ノードに解決する (明示接続 > 幾何近接)。"""
    if cxn_id and cxn_id in node_by_id:
        return node_by_id[cxn_id]
    best, best_dist = None, None
    for n in nodes:
        d = _point_rect_distance(point, n["rect"])
        if best_dist is None or d < best_dist:
            best, best_dist = n, d
    # 近接が遠すぎる (端点が図形に接していない) 場合は誤接続を避けて棄却。
    if best_dist is not None and best_dist > _CXN_SNAP:
        return None
    return best


def _point_rect_distance(point: tuple[int, int], rect: tuple[int, int, int, int]) -> int:
    """点 (col,row) と矩形 (col0,row0,col1,row1) のマンハッタン距離 (内側は 0)。"""
    px, py = point
    c0, r0, c1, r1 = rect
    dx = max(c0 - px, px - c1, 0)
    dy = max(r0 - py, py - r1, 0)
    return dx + dy


def _cxn_end_id(cxn, tag: str) -> str | None:
    """コネクタの端点接続 (<a:stCxn>/<a:endCxn>) から接続先シェイプ id を返す。"""
    el = cxn.find(".//{%s}%s" % (_NS["a"], tag))
    return el.get("id") if el is not None else None


def _anchor_rect(anchor) -> tuple[int, int, int, int]:
    """アンカーの from/to からセル座標の外接矩形 (col0,row0,col1,row1) を作る。

    to が無い (oneCellAnchor 等) 場合は from を点として扱う。値が読めない
    場合は原点 (0,0,0,0)。
    """
    c0, r0 = _anchor_point(anchor, "from")
    c1, r1 = _anchor_point(anchor, "to")
    if c1 is None or r1 is None:
        c1, r1 = c0, r0
    return (c0 or 0, r0 or 0, c1 or 0, r1 or 0)


def _anchor_point(anchor, tag: str) -> tuple[int | None, int | None]:
    node = anchor.find("{%s}%s" % (_NS["xdr"], tag))
    if node is None:
        return (None, None)
    col = node.findtext("{%s}col" % _NS["xdr"])
    row = node.findtext("{%s}row" % _NS["xdr"])
    try:
        return (int(col), int(row)) if col is not None and row is not None else (None, None)
    except ValueError:
        return (None, None)


def _rect_cell(rect: tuple[int, int, int, int]) -> str | None:
    """矩形左上をアンカーセル (A1 形式) に。"""
    try:
        return f"{_col_letter(rect[0] + 1)}{rect[1] + 1}"
    except (ValueError, TypeError):
        return None


def _shape_id(sp) -> str:
    cnvpr = sp.find(".//{%s}cNvPr" % _NS["xdr"])
    return cnvpr.get("id", "") if cnvpr is not None else ""


def _shape_name(sp) -> str:
    cnvpr = sp.find(".//{%s}cNvPr" % _NS["xdr"])
    return cnvpr.get("name", "") if cnvpr is not None else ""


def _shape_text(sp) -> str:
    """図形の txBody から段落テキストを組み立てる (段落=改行区切り)。"""
    tx = sp.find("{%s}txBody" % _NS["xdr"])
    if tx is None:
        return ""
    lines = []
    for p in tx.findall("{%s}p" % _NS["a"]):
        runs = "".join(t.text or "" for t in p.findall("{%s}r/{%s}t" % (_NS["a"], _NS["a"])))
        if runs.strip():
            lines.append(runs)
    return "\n".join(lines).strip()


_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
