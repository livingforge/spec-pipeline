"""Excel (.xlsx) の抽出器。

各シートの使用範囲を 1 つの表として抽出する(末尾の空行・空列は除去)。
シートに埋め込まれた画像も取り出す。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..models import ExtractionResult, ImageElement, TableElement
from .base import ImageSaver


def extract_xlsx(path: Path, saver: ImageSaver) -> ExtractionResult:
    # data_only=True で数式ではなくキャッシュされた計算結果を読む
    wb = load_workbook(str(path), data_only=True)
    result = ExtractionResult(source=path.name, file_type="xlsx")

    props = wb.properties
    result.metadata = {
        "title": props.title or None,
        "author": props.creator or None,
        "created": props.created.isoformat() if props.created else None,
        "modified": props.modified.isoformat() if props.modified else None,
        "sheets": wb.sheetnames,
    }

    for ws in wb.worksheets:
        rows = _sheet_to_rows(ws)
        if rows:
            result.elements.append(
                TableElement(rows=rows, location={"sheet": ws.title})
            )
        for img in getattr(ws, "_images", []):
            try:
                data = img._data()
            except Exception:
                continue
            ext = getattr(img, "format", None) or "png"
            rel_path = saver.save(data, str(ext))
            result.elements.append(
                ImageElement(
                    file=rel_path,
                    format=str(ext).lower(),
                    location={"sheet": ws.title, "anchor": _anchor_cell(img)},
                )
            )
    return result


def _sheet_to_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    # 末尾の空行を除去
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    # 末尾の空列を除去
    while rows and all(r and r[-1] == "" for r in rows):
        for r in rows:
            r.pop()
    return rows


def _anchor_cell(img) -> str | None:
    try:
        anc = img.anchor._from
        # 0 始まりの列番号を A1 形式に変換
        col = ""
        n = anc.col + 1
        while n:
            n, rem = divmod(n - 1, 26)
            col = chr(65 + rem) + col
        return f"{col}{anc.row + 1}"
    except Exception:
        return None
